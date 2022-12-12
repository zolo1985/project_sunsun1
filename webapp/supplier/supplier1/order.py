from io import BytesIO
from flask import (Blueprint, flash, redirect, render_template, request,
                   url_for, abort, send_file, jsonify)
from webapp import has_role
from flask_login import current_user, login_required
from webapp import models
from webapp.database import Connection
from webapp.supplier.supplier1.forms import OrderAddForm, FiltersForm, OrderDetailFileAddForm, SubmitFileOrders
from datetime import datetime, time, timedelta
from sqlalchemy import func, or_
import logging
from webapp.utils import is_time_between
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import pytz

log = logging.getLogger(__name__)

supplier1_order_blueprint = Blueprint('supplier1_order', __name__)

@supplier1_order_blueprint.route('/supplier1/orders', methods=['GET', 'POST'])
@login_required
@has_role('supplier1')
def supplier1_orders():
    connection = Connection()
    cur_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
    form = FiltersForm()
    orders1 = connection.query(models.Delivery).filter(models.Delivery.user_id==current_user.id).filter(models.Delivery.user_id==current_user.id).filter(func.date(models.Delivery.created_date) == cur_date.date()).all()
    post_orders = connection.query(models.Delivery).filter(models.Delivery.user_id==current_user.id).filter(func.date(models.Delivery.postphoned_date) == cur_date.date()).filter(models.Delivery.is_postphoned == True).all()
    orders = orders1 + post_orders

    if form.validate_on_submit():
        if form.date.data is not None:
            orders1 = connection.query(models.Delivery).filter(func.date(models.Delivery.created_date) == form.date.data).filter(models.Delivery.user_id==current_user.id).all()
            post_orders = connection.query(models.Delivery).filter(models.Delivery.user_id==current_user.id).filter(func.date(models.Delivery.postphoned_date) == form.date.data).filter(models.Delivery.is_postphoned == True).all()

            for post_order in post_orders:
                if post_order in orders1:
                    continue
                else:
                    orders1.append(post_order)
            orders = orders1
            
            return render_template('/supplier/supplier1/orders.html', orders=orders, cur_date=cur_date, form=form)

    return render_template('/supplier/supplier1/orders.html', orders=orders, cur_date=cur_date, form=form)


@supplier1_order_blueprint.route('/supplier1/orders/add', methods=['GET', 'POST'])
@login_required
@has_role('supplier1')
def supplier1_order_add():
    cur_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")).date()
    order_window = is_time_between(time(22,30), time(00,00))

    connection = Connection()
    districts = connection.query(models.District).all()
    aimags = connection.query(models.Aimag).all()
    last_five_orders = connection.query(models.Delivery).filter(models.Delivery.user_id==current_user.id).order_by(models.Delivery.created_date.desc()).limit(5)

    form = OrderAddForm()

    form.district.choices = [(district) for district in districts]
    form.district.choices.insert(0,'Дүүрэг сонгох')
    form.khoroo.choices = [(f'%s'%(district+1)) for district in range(32)]
    form.khoroo.choices.insert(0,'Хороо сонгох')
    form.aimag.choices = [(aimag) for aimag in aimags]
    form.aimag.choices.insert(0,'Аймаг сонгох')

    if form.validate_on_submit():
        line_products = request.form.getlist("product")
        line_quantities = request.form.getlist("quantity")

        if form.order_type.data == "0":
            for i, product in enumerate(line_products):
                quantity = connection.execute('SELECT SUM(quantity) as quantity FROM sunsundatabase1.total_inventory inventory join sunsundatabase1.product product on product.id=inventory.product_id where product.supplier_id=:current_user and product.id=:product_id group by product_id;', {'current_user': current_user.id, 'product_id': int(line_products[i])}).scalar()
                    
                if int(quantity) < int(line_quantities[i]):
                    flash(f'Агуулахад байгаа барааны тоо хүрэхгүй байна', 'danger')
                    return redirect(url_for('supplier1_order.supplier1_order_add'))

            order = models.Delivery()
            order.status = "unassigned"
            order.destination_type = "local"
            order.order_type = "stored"
            order.is_ready = True
            order.delivery_attempts = 0
            order.supplier_company_name = current_user.company_name
            order.total_amount = abs(form.total_amount.data)

            if is_time_between(time(22,30), time(00,00)):
                order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
            else:
                order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

            current_user.deliveries.append(order)
            connection.flush()

            address = models.Address()
            address.phone = form.phone.data
            address.phone_more = form.phone_more.data
            address.district = form.district.data
            address.khoroo = form.khoroo.data
            address.address = form.address.data
            address.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
            address.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

            order.addresses = address

            for i, product in enumerate(line_products):
                order_detail = models.DeliveryDetail()
                order_detail.quantity = int(line_quantities[i])
                order_detail.product_id = int(line_products[i])
                order_detail.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                order_detail.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                total_inventory_product = connection.query(models.TotalInventory).filter_by(product_id=int(line_products[i])).first()
                total_inventory_product.quantity = total_inventory_product.quantity-int(line_quantities[i])
                total_inventory_product.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                is_detail = connection.query(models.DeliveryDetail).filter(models.DeliveryDetail.delivery_id==order.id, models.DeliveryDetail.product_id==int(line_products[i])).first()

                if is_detail:
                    is_detail.quantity = is_detail.quantity + int(line_quantities[i])
                else:
                    order.delivery_details.append(order_detail)
                    connection.flush()
                
            try:
                connection.commit()
            except Exception as ex:
                flash('Алдаа гарлаа!', 'danger')
                connection.rollback()
                return redirect(url_for('supplier1_order.supplier1_orders'))
            else:
                if is_time_between(time(22,30), time(00,00)):
                    flash('Маргаашийн хүргэлтэнд нэмэгдлээ.', 'success')
                else:
                    flash('Хүргэлт нэмэгдлээ.', 'success')
                return redirect(url_for('supplier1_order.supplier1_order_add', destination='local'))

        elif form.order_type.data == "1":
            for i, product in enumerate(line_products):
                quantity = connection.execute('SELECT SUM(quantity) as quantity FROM sunsundatabase1.total_inventory inventory join sunsundatabase1.product product on product.id=inventory.product_id where product.supplier_id=:current_user and product.id=:product_id group by product_id;', {'current_user': current_user.id, 'product_id': int(line_products[i])}).scalar()
                    
                if int(quantity) < int(line_quantities[i]):
                    flash(f'Агуулахад байгаа барааны тоо хүрэхгүй байна', 'danger')
                    return redirect(url_for('supplier1_order.supplier1_order_add'))
            
            order = models.Delivery()
            order.status = "unassigned"
            order.destination_type = "long"
            order.order_type = "stored"
            order.is_ready = True
            order.delivery_attempts = 0
            order.supplier_company_name = current_user.company_name
            order.total_amount = form.total_amount.data

            if is_time_between(time(22,30), time(00,00)):
                order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
            else:
                order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

            current_user.deliveries.append(order)
            connection.flush()
            
            address = models.Address()
            address.phone = form.phone.data
            address.phone_more = form.phone_more.data
            address.aimag = form.aimag.data
            address.address = form.address.data
            address.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
            address.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
            
            order.addresses = address
            
            for i, product in enumerate(line_products):
                order_detail = models.DeliveryDetail()
                order_detail.quantity = int(line_quantities[i])
                order_detail.product_id = int(line_products[i])
                order_detail.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                order_detail.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                total_inventory_product = connection.query(models.TotalInventory).filter_by(product_id=int(line_products[i])).first()
                total_inventory_product.quantity = total_inventory_product.quantity-int(line_quantities[i])
                total_inventory_product.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                is_detail = connection.query(models.DeliveryDetail).filter(models.DeliveryDetail.delivery_id==order.id, models.DeliveryDetail.product_id==int(line_products[i])).first()

                if is_detail:
                    is_detail.quantity = is_detail.quantity + int(line_quantities[i])
                else:
                    order.delivery_details.append(order_detail)
                    connection.flush()
                
            try:
                connection.commit()
            except Exception as ex:
                flash('Алдаа гарлаа!', 'danger')
                connection.rollback()
                return redirect(url_for('supplier1_order.supplier1_orders'))
            else:
                if is_time_between(time(22,30), time(00,00)):
                    flash('Маргаашийн хүргэлтэнд нэмэгдлээ.', 'success')
                else:
                    flash('Хүргэлт нэмэгдлээ.', 'success')
                return redirect(url_for('supplier1_order.supplier1_order_add', destination='long'))
    return render_template('/supplier/supplier1/order_add.html', cur_date=cur_date, order_window=order_window, form=form, last_five_orders=last_five_orders)



@supplier1_order_blueprint.route('/supplier1/orders/add/excel', methods=['GET', 'POST'])
@login_required
@has_role('supplier1')
def supplier1_order_add_excel():
    cur_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")).date()
    order_window = is_time_between(time(22,30), time(00,00))
    form2 = OrderDetailFileAddForm()
    form3 = SubmitFileOrders()
    orders = []
    insufficient_product_ids = []
    products_list = []
    products_list_result = []

    if form2.validate_on_submit():
        orders = []
        insufficient_product_ids = []
        products_list = []
        products_list_result = []
        connection = Connection()
        if form2.validate_on_submit():
            if len(request.files) > 0 and request.files['excel_file'].filename != "":
                excel_file = request.files['excel_file']
                df = pd.read_excel(excel_file)

                phones = df['Утасны дугаар'].values
                districts = df['Дүүрэг'].values
                khoroos = df['Хороо'].values
                aimags = df['Аймаг'].values
                addresses = df['Хаяг'].values
                quantities = df['Тоо ширхэг'].values
                products = df['Бараа'].values
                total_amounts = df['Нийт үнэ'].values

                for i, phone in enumerate(phones):
                    if (str(districts[i]) != 'nan') and (str(khoroos[i]) != 'nan'):
                        order_dict = {
                            "Захиалгын төрөл": "local",
                            "Утасны дугаар": phone,
                            "Дүүрэг": districts[i],
                            "Хороо": int(khoroos[i]),
                            "Хаяг": addresses[i],
                            "Бараа": [f"%s, %s"%(products[i], int(abs(quantities[i])))],
                            "Нийт үнэ": abs(total_amounts[i])
                        }
                        orders.append(order_dict)
                    else:
                        order_dict = {
                            "Захиалгын төрөл": "long",
                            "Утасны дугаар": phone,
                            "Аймаг": aimags[i],
                            "Хаяг": addresses[i],
                            "Бараа": [f"%s, %s"%(products[i], int(abs(quantities[i])))],
                            "Нийт үнэ": abs(total_amounts[i])
                        }
                        orders.append(order_dict)

                for i, order in enumerate(orders):
                    if i == len(orders) - 1:
                        continue
                    else:
                        if order["Утасны дугаар"] == orders[i+1]["Утасны дугаар"] and order["Хаяг"] == orders[i+1]["Хаяг"]:
                            order["Бараа"].append(orders[i+1]["Бараа"][0])
                            order["Нийт үнэ"] = order["Нийт үнэ"] + orders[i+1]["Нийт үнэ"]
                            orders.pop(i+1)

                for i, product in enumerate(products):
                    
                    product_dict = {
                        "product_id": product.split(".", 1)[0],
                        "quantity": quantities[i]
                    }

                    products_list.append(product_dict)

                for i, orig_obj in enumerate(products_list):
                    if i==0:
                        products_list_result.append(orig_obj)
                    else:
                        if products_list_result[-1]["product_id"]==orig_obj["product_id"]:
                            products_list_result[-1]["quantity"]=products_list_result[-1]["quantity"] + orig_obj["quantity"]
                        else:
                            products_list_result.append(orig_obj)

                for i, product_sum in enumerate(products_list_result):
                    invent_quantity = connection.query(models.TotalInventory.quantity).filter_by(product_id=int(product_sum["product_id"])).scalar()
                    postphoned_invent_quantity = connection.query(models.TotalInventory.postphoned_quantity).filter_by(product_id=int(product_sum["product_id"])).scalar()
                    cancelled_invent_quantity = connection.query(models.TotalInventory.cancelled_quantity).filter_by(product_id=int(product_sum["product_id"])).scalar()

                    total_invent_quantity = int(invent_quantity) + int(postphoned_invent_quantity) + (cancelled_invent_quantity)

                    if product_sum["quantity"]>total_invent_quantity:
                        insufficient_product_ids.append(product_sum["product_id"])

                if len(insufficient_product_ids) > 0:
                    flash('Сүнсүн агуулахад байгаа барааг хүрэлцэхгүй байна!', 'danger')
                    return render_template('/supplier/supplier1/order_add_excel.html', cur_date=cur_date, form2=form2, form3=form3, orders=orders, insufficient_product_ids=insufficient_product_ids, order_window=order_window)
                else:
                    for i, order in enumerate(orders):
                        if order["Захиалгын төрөл"] == "local":
                            new_order = models.Delivery()
                            new_order.status = "unassigned"
                            new_order.destination_type = "local"
                            new_order.order_type = "stored"
                            new_order.is_ready = True
                            new_order.delivery_attempts = 0
                            new_order.supplier_company_name = current_user.company_name
                            new_order.total_amount = order["Нийт үнэ"]

                            if is_time_between(time(22,30), time(00,00)):
                                new_order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                                new_order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                                new_order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                            else:
                                new_order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                                new_order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                                new_order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                            current_user.deliveries.append(new_order)

                            address = models.Address()
                            address.phone = order["Утасны дугаар"]
                            address.district = order["Дүүрэг"]
                            address.khoroo = order["Хороо"]
                            address.address = order["Хаяг"]
                            address.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                            address.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                            new_order.addresses = address

                            for i, product in enumerate(order["Бараа"]):
                                
                                order_detail = models.DeliveryDetail()
                                order_detail.quantity = int(product.split(",", 1)[1])
                                order_detail.product_id = int(product.split(".", 1)[0])
                                order_detail.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                                order_detail.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                                total_inventory_product = connection.query(models.TotalInventory).filter_by(product_id=int(product.split(".", 1)[0])).first()
                                total_inventory_product.quantity = total_inventory_product.quantity-int(product.split(",", 1)[1])
                                total_inventory_product.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                                new_order.delivery_details.append(order_detail)
                            
                        elif order["Захиалгын төрөл"] == "long":
                            new_order = models.Delivery()
                            new_order.status = "unassigned"
                            new_order.destination_type = "long"
                            new_order.order_type = "stored"
                            new_order.is_ready = True
                            new_order.delivery_attempts = 0
                            new_order.supplier_company_name = current_user.company_name
                            new_order.total_amount = order["Нийт үнэ"]

                            if is_time_between(time(22,30), time(00,00)):
                                new_order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                                new_order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                                new_order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar")) + timedelta(hours=+24)
                            else:
                                new_order.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                                new_order.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                                new_order.delivery_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                            current_user.deliveries.append(new_order)

                            address = models.Address()
                            address.phone = order["Утасны дугаар"]
                            address.aimag = order["Аймаг"]
                            address.address = order["Хаяг"]
                            address.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                            address.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                            new_order.addresses = address

                            for i, product in enumerate(order["Бараа"]):
                                
                                order_detail = models.DeliveryDetail()
                                order_detail.quantity = int(product.split(",", 1)[1])
                                order_detail.product_id = int(product.split(".", 1)[0])
                                order_detail.created_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                                order_detail.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))

                                total_inventory_product = connection.query(models.TotalInventory).filter_by(product_id=int(product.split(".", 1)[0])).first()
                                total_inventory_product.quantity = total_inventory_product.quantity-int(product.split(",", 1)[1])
                                total_inventory_product.modified_date = datetime.now(pytz.timezone("Asia/Ulaanbaatar"))
                                new_order.delivery_details.append(order_detail)
                        else:
                            flash('Захиалга алдаатай дахин шалгана уу!', 'danger')
                            return render_template('/supplier/supplier1/order_add_excel.html', cur_date=cur_date, form2=form2, form3=form3, orders=orders, insufficient_product_ids=insufficient_product_ids, order_window=order_window)

                    try:
                        connection.commit()
                    except Exception:
                        flash('Алдаа гарлаа!', 'danger')
                        connection.rollback()
                    else:
                        if is_time_between(time(22,30), time(00,00)):
                            flash('Маргаашийн хүргэлтэнд нэмэгдлээ.', 'success')
                        else:
                            flash('Хүргэлт нэмэгдлээ.', 'success')
                        return redirect(url_for('supplier1_order.supplier1_orders'))

    return render_template('/supplier/supplier1/order_add_excel.html', cur_date=cur_date, order_window=order_window, form2=form2, form3=form3, orders=orders, insufficient_product_ids=insufficient_product_ids)



@supplier1_order_blueprint.route('/supplier1/orders/add/file/template/download', methods=['GET', 'POST'])
@login_required
@has_role('supplier1')
def supplier1_order_excel_template():
    wb = Workbook()
    ws2 = wb.create_sheet("Бараа")
    ws = wb.active
    ws.title = "Хүргэлт"

    ws.append(['Утасны дугаар', 'Дүүрэг', 'Хороо', 'Аймаг', 'Хаяг', 'Бараа', 'Тоо ширхэг', 'Нийт үнэ'])

    dv = DataValidation(type="list", formula1='Бараа!$B$1:$B$1000', allow_blank=False)
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'

    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'List Selection'

    dv1 = DataValidation(type="list", formula1='Бараа!$G$1:$G$1000', allow_blank=False)
    dv2 = DataValidation(type="whole", allow_blank=False, operator="greaterThanOrEqual", formula1=0)
    dv3 = DataValidation(type="list", formula1='Бараа!$M$1:$M$1000', allow_blank=False)
    dv4 = DataValidation(type="list", formula1='Бараа!$K$1:$K$1000', allow_blank=False)

    ws.add_data_validation(dv)
    ws.add_data_validation(dv1)
    ws.add_data_validation(dv2)
    ws.add_data_validation(dv3)
    ws.add_data_validation(dv4)

    dv.add("F2:F1000")
    dv1.add("B2:B1000")
    dv2.add("H2:H1000")
    dv2.add("A2:A1000")
    dv2.add("G2:G1000")
    dv3.add("C2:C1000")
    dv4.add("D2:D1000")

    connection = Connection()
    products = connection.query(models.Product).join(models.TotalInventory).filter(models.Product.supplier_id == current_user.id).filter(models.TotalInventory.quantity > 0).all()
    districts = connection.query(models.District).all()
    aimags = connection.query(models.Aimag).all()

    for i in range(1, len(products)+1):
        ws2.cell(row=i, column=1).value = products[i-1].id
        ws2.cell(row=i, column=2).value = f"%s. %s. %s. %s. ₮%s"%(products[i-1].id, products[i-1].name, str(products[i-1].colors[0]), str(products[i-1].sizes[0]), products[i-1].price)
        ws2.cell(row=i, column=3).value = products[i-1].price
        ws2.cell(row=i, column=4).value = str(products[i-1].colors[0])
        ws2.cell(row=i, column=5).value = str(products[i-1].sizes[0])

    for i in range(1, len(districts)+1):
        ws2.cell(row=i, column=7).value = districts[i-1].name

    for i in range(1, len(aimags)+1):
        ws2.cell(row=i, column=11).value = aimags[i-1].name

    for i in range(1, 50):
        ws2.cell(row=i, column=13).value = i

    file_stream = BytesIO()
    wb.worksheets[1].sheet_state = "hidden"
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(file_stream, attachment_filename="order_template.xlsx", as_attachment=True)



@supplier1_order_blueprint.route("/supplier1/orders/search/products", methods=['GET', 'POST'])
@login_required
@has_role('supplier1')
def supplier1_search_products():
    query = request.form.get("term")
    connection = Connection()
    search = "%{}%".format(query)
    products = connection.query(models.Product).filter(models.Product.supplier_id==current_user.id, models.Product.name.like(search)).all()
    results = []
    for product in products:
        results.append({
            'id': product.id,
            'name': product.name,
            'color': str(product.colors[0]),
            'size': str(product.sizes[0]),
            'quantity': int(connection.query(models.TotalInventory.quantity).filter(models.TotalInventory.product_id==product.id).scalar()),
            'price': int(product.price),
        })
    return jsonify(results)



@supplier1_order_blueprint.route('/supplier1/orders/delete/<int:order_id>', methods=['GET', 'POST'])
@login_required
@has_role('supplier1')
def supplier1_order_delete(order_id):
    connection = Connection()
    order_to_delete = connection.query(models.Delivery).filter(models.Delivery.id==order_id).first()

    if order_to_delete is None:
        flash('Захиалга олдсонгүй!', 'danger')
        connection.close()
        return redirect(url_for('supplier1_order.supplier1_orders'))

    if order_to_delete.user_id != current_user.id:
        connection.close()
        abort(403)

    if order_to_delete.status != "unassigned" or order_to_delete.delivery_region is not None:
        flash('Захиалга цуцлах боломжгүй байна! Менежертэй холбоо барина уу!', 'danger')
        connection.close()
        return redirect(url_for('supplier1_order.supplier1_orders'))
    else:
        
        for order_detail in order_to_delete.delivery_details:
            total_product_inventory = connection.query(models.TotalInventory).filter(models.TotalInventory.product_id==order_detail.product_id).first()
            total_product_inventory.quantity = total_product_inventory.quantity + order_detail.quantity
        
        connection.query(models.Address).filter_by(delivery_id=order_to_delete.id).delete()
        connection.query(models.DeliveryDetail).filter_by(delivery_id=order_to_delete.id).delete()
        connection.query(models.Delivery).filter_by(id=order_to_delete.id).delete()

        try:
            connection.commit()
        except Exception as ex:
            flash(str(ex), 'danger')
            connection.rollback()
            connection.close()
            return redirect(url_for('supplier1_order.supplier1_orders'))
        else:
            flash('Устгалаа', 'success')
            return redirect(url_for('supplier1_order.supplier1_orders'))
    